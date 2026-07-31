from __future__ import annotations

import csv
import io
from dataclasses import FrozenInstanceError

import pytest
from openpyxl import Workbook

from pallet_optimizer.domain import CargoItem, DomainError, Margins, Shape
from pallet_optimizer.normalization import normalize_payload, payload_from_csv, payload_from_xlsx


def test_domain_is_immutable_and_rejects_impossible_state() -> None:
    item = CargoItem("x", "x", 0, Shape.PALLET, 1200, 800, 1000, 100, "A", 1)
    with pytest.raises(FrozenInstanceError):
        item.length_mm = 1  # type: ignore[misc]
    with pytest.raises(DomainError) as error:
        CargoItem("bad", "bad", 0, Shape.PALLET, 0, 800, 1000, 100, "A", 1)
    assert error.value.diagnostic.code == "INVALID_DIMENSION"


def test_units_quantities_and_margins_are_normalized() -> None:
    problem = normalize_payload({
        "dimension_unit": "m", "weight_unit": "t", "default_margins": {"left": 0.01, "right": 0.02},
        "items": [{"id": "P", "quantity": 2, "shape": "irregular", "length": 1.2, "width": 0.8,
                   "height": 1, "weight": 0.5, "destination": "A", "delivery_order": 1}],
    })
    assert len(problem.items) == 2
    assert problem.items[0].length_mm == 1200
    assert problem.items[0].weight_kg == 500
    assert problem.items[0].margins == Margins(left_mm=10, right_mm=20)


def test_csv_xlsx_and_json_produce_identical_problem() -> None:
    row = {"id": "P1", "quantity": 2, "shape": "pallet", "length": 1200, "width": 800,
           "height": 1000, "weight": 500, "destination": "A", "delivery_order": 1,
           "rotation_allowed": True}
    base = {"vehicle_policy": {"mode": "forced", "forced_vehicle_id": "semi_trailer"}}
    expected = normalize_payload({**base, "items": [row]})
    text = io.StringIO()
    writer = csv.DictWriter(text, fieldnames=row.keys(), delimiter=";")
    writer.writeheader(); writer.writerow(row)
    from_csv = normalize_payload(payload_from_csv(text.getvalue().encode(), **base))
    wb = Workbook(); ws = wb.active; ws.append(list(row)); ws.append(list(row.values()))
    binary = io.BytesIO(); wb.save(binary)
    from_xlsx = normalize_payload(payload_from_xlsx(binary.getvalue(), **base))
    assert from_csv == expected
    assert from_xlsx == expected


def test_xlsx_with_meter_headers_is_detected_and_normalized() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import total"
    sheet.append(["Référence", "Qté", "Forme", "L (m)", "l (m)", "H (m)", "Poids (kg)", "Destination"])
    sheet.append(["PAL-M", 1, "pallet", 1.2, 0.8, 1.4, 500, "Client A"])
    binary = io.BytesIO(); workbook.save(binary)

    payload = payload_from_xlsx(binary.getvalue())
    problem = normalize_payload(payload)

    assert payload["dimension_unit"] == "m"
    assert payload["_import_sheet"] == "Import total"
    assert problem.items[0].length_mm == 1200
    assert problem.items[0].width_mm == 800
    assert problem.items[0].height_mm == 1400


def test_xlsx_missing_columns_returns_a_precise_message() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Référence", "Qté", "Poids (kg)"])
    sheet.append(["PAL-1", 1, 500])
    binary = io.BytesIO(); workbook.save(binary)

    with pytest.raises(DomainError) as error:
        payload_from_xlsx(binary.getvalue())

    assert error.value.diagnostic.code == "MISSING_COLUMNS"
    assert "Longueur" in error.value.diagnostic.message
    assert "Largeur" in error.value.diagnostic.message
    assert "Hauteur" in error.value.diagnostic.message


def test_corrupted_xlsx_returns_a_conversion_instruction() -> None:
    with pytest.raises(DomainError) as error:
        payload_from_xlsx(b"not-an-excel-workbook")

    assert error.value.diagnostic.code == "INVALID_XLSX_FILE"
    assert "Enregistrer sous" in error.value.diagnostic.message
    assert ".xlsx" in error.value.diagnostic.message


def test_invalid_import_has_structured_field_path() -> None:
    with pytest.raises(DomainError) as error:
        normalize_payload({"items": [{"id": "X", "length": "oops", "width": 1, "height": 1, "weight": 1}]})
    assert error.value.diagnostic.code == "INVALID_NUMBER"
    assert "length" in (error.value.diagnostic.field_path or "")
    assert "n’est pas un nombre valide" in error.value.diagnostic.message


def test_maximum_100_expanded_objects() -> None:
    payload = {"items": [{"id": "X", "quantity": 101, "length": 100, "width": 100, "height": 100, "weight": 1}]}
    with pytest.raises(DomainError) as error:
        normalize_payload(payload)
    assert error.value.diagnostic.code == "TOO_MANY_ITEMS"
