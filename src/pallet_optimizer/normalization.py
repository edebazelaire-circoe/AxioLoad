from __future__ import annotations

import csv
import io
import re
import unicodedata
from collections.abc import Mapping
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .catalog import default_vehicle_catalog, find_vehicle
from .domain import CargoItem, Diagnostic, DomainError, Margins, OptimizationProblem, Shape, VehiclePolicy, VehicleVersion

_DIMENSION_FACTORS = {"mm": 1.0, "cm": 10.0, "m": 1000.0}
_WEIGHT_FACTORS = {"kg": 1.0, "g": 0.001, "t": 1000.0}
_REQUIRED_IMPORT_COLUMNS = {"length", "width", "height", "weight"}
_FRIENDLY_FIELDS = {
    "quantity": "Qté",
    "length": "Longueur",
    "width": "Largeur",
    "height": "Hauteur",
    "weight": "Poids",
    "delivery_order": "Ordre",
    "separation": "Séparation",
}


def _number(value: Any, field: str, *, label: str | None = None) -> float:
    displayed = label or field
    if value is None or str(value).strip() == "":
        raise DomainError(
            Diagnostic(
                "MISSING_VALUE",
                f"{displayed} : valeur obligatoire manquante. La cellule est vide ou contient une formule Excel non calculée.",
                field_path=field,
            )
        )
    try:
        return float(str(value).strip().replace(" ", "").replace(",", "."))
    except (TypeError, ValueError) as exc:
        raise DomainError(
            Diagnostic(
                "INVALID_NUMBER",
                f"{displayed} : « {value} » n’est pas un nombre valide. Utilisez uniquement des chiffres, avec une virgule ou un point pour les décimales.",
                field_path=field,
            )
        ) from exc


def _bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "oui", "y", "o"}


def _tokens(value: Any) -> tuple[str, ...]:
    if value is None or value == "":
        return ()
    if isinstance(value, (list, tuple)):
        return tuple(str(v).strip() for v in value if str(v).strip())
    return tuple(part.strip() for part in str(value).replace(";", ",").split(",") if part.strip())


def _field_label(raw: Mapping[str, Any], input_index: int, field: str) -> str:
    row_number = raw.get("_source_row")
    friendly = _FRIENDLY_FIELDS.get(field, field)
    if row_number:
        return f"Ligne {row_number}, colonne « {friendly} »"
    return f"Élément {input_index + 1}, champ « {friendly} »"


def normalize_payload(
    payload: Mapping[str, Any],
    *,
    requested_solutions: int | None = None,
    catalog: tuple[VehicleVersion, ...] | None = None,
) -> OptimizationProblem:
    raw_items = payload.get("items")
    if not isinstance(raw_items, list) or not raw_items:
        raise DomainError(
            Diagnostic(
                "INVALID_ITEMS",
                "Le fichier ne contient aucune ligne de marchandise exploitable. Vérifiez que les données se trouvent sous la ligne d’en-tête.",
                field_path="items",
            )
        )
    default_dimension_unit = str(payload.get("dimension_unit", "mm")).lower()
    default_weight_unit = str(payload.get("weight_unit", "kg")).lower()
    if default_dimension_unit not in _DIMENSION_FACTORS or default_weight_unit not in _WEIGHT_FACTORS:
        raise DomainError(
            Diagnostic(
                "INVALID_UNIT",
                "Unité non reconnue. Les dimensions acceptées sont mm, cm ou m et les poids g, kg ou t.",
            )
        )
    defaults = payload.get("default_margins", {}) or {}
    expanded: list[CargoItem] = []
    for input_index, raw in enumerate(raw_items):
        if not isinstance(raw, Mapping):
            raise DomainError(
                Diagnostic(
                    "INVALID_ITEM",
                    f"La ligne {input_index + 2} du fichier n’a pas pu être interprétée.",
                    field_path=f"items[{input_index}]",
                )
            )
        dim_unit = str(raw.get("dimension_unit", default_dimension_unit)).lower()
        weight_unit = str(raw.get("weight_unit", default_weight_unit)).lower()
        if dim_unit not in _DIMENSION_FACTORS or weight_unit not in _WEIGHT_FACTORS:
            raise DomainError(
                Diagnostic(
                    "INVALID_UNIT",
                    f"{_field_label(raw, input_index, 'dimension_unit')} : unité invalide. Utilisez mm, cm ou m.",
                    field_path=f"items[{input_index}].dimension_unit",
                )
            )
        df, wf = _DIMENSION_FACTORS[dim_unit], _WEIGHT_FACTORS[weight_unit]
        source_id = str(raw.get("id") or f"item-{input_index + 1}")
        quantity = int(
            _number(
                raw.get("quantity", 1),
                f"items[{input_index}].quantity",
                label=_field_label(raw, input_index, "quantity"),
            )
        )
        if quantity < 1:
            raise DomainError(
                Diagnostic(
                    "INVALID_QUANTITY",
                    f"{_field_label(raw, input_index, 'quantity')} : la quantité doit être au moins égale à 1.",
                    field_path=f"items[{input_index}].quantity",
                )
            )
        raw_margins = raw.get("margins", {}) or {}

        def margin(name: str) -> int:
            return round(
                _number(
                    raw_margins.get(name, defaults.get(name, 0)),
                    f"items[{input_index}].margins.{name}",
                    label=f"Marge {name} de l’élément {input_index + 1}",
                )
                * df
            )

        margins = Margins(margin("left"), margin("right"), margin("front"), margin("rear"), margin("top"))
        shape_value = str(raw.get("shape", "pallet")).strip().lower()
        try:
            shape = Shape(shape_value)
        except ValueError as exc:
            row_number = raw.get("_source_row", input_index + 2)
            raise DomainError(
                Diagnostic(
                    "INVALID_SHAPE",
                    f"Ligne {row_number}, colonne « Forme » : « {shape_value} » n’est pas reconnu. Utilisez pallet, box, roll, cylinder, sheet, post, bar_rect ou bar_cyl.",
                    field_path=f"items[{input_index}].shape",
                )
            ) from exc
        for copy_index in range(quantity):
            item_id = source_id if quantity == 1 else f"{source_id}#{copy_index + 1}"
            expanded.append(
                CargoItem(
                    id=item_id,
                    source_id=source_id,
                    input_index=input_index,
                    shape=shape,
                    length_mm=round(
                        _number(
                            raw.get("length"),
                            f"items[{input_index}].length",
                            label=_field_label(raw, input_index, "length"),
                        )
                        * df
                    ),
                    width_mm=round(
                        _number(
                            raw.get("width"),
                            f"items[{input_index}].width",
                            label=_field_label(raw, input_index, "width"),
                        )
                        * df
                    ),
                    height_mm=round(
                        _number(
                            raw.get("height"),
                            f"items[{input_index}].height",
                            label=_field_label(raw, input_index, "height"),
                        )
                        * df
                    ),
                    weight_kg=_number(
                        raw.get("weight"),
                        f"items[{input_index}].weight",
                        label=_field_label(raw, input_index, "weight"),
                    )
                    * wf,
                    destination=str(raw.get("destination") or f"Destination {input_index + 1}"),
                    delivery_order=int(
                        _number(
                            raw.get("delivery_order", input_index + 1),
                            f"items[{input_index}].delivery_order",
                            label=_field_label(raw, input_index, "delivery_order"),
                        )
                    ),
                    rotation_allowed=_bool(raw.get("rotation_allowed"), True),
                    stackable=_bool(raw.get("stackable"), False),
                    margins=margins,
                    compatibility_tags=_tokens(raw.get("compatibility_tags")),
                    incompatible_tags=_tokens(raw.get("incompatible_tags")),
                    keep_together_group=str(raw["keep_together_group"]) if raw.get("keep_together_group") else None,
                    separate_group=str(raw["separate_group"]) if raw.get("separate_group") else None,
                    separation_mm=round(
                        _number(
                            raw.get("separation", 0),
                            f"items[{input_index}].separation",
                            label=_field_label(raw, input_index, "separation"),
                        )
                        * df
                    ),
                    zone=str(raw["zone"]) if raw.get("zone") else None,
                )
            )
    catalog = catalog or default_vehicle_catalog()
    policy_raw = payload.get("vehicle_policy", {}) or {}
    policy = VehiclePolicy(
        mode=str(policy_raw.get("mode", "auto")),
        forced_vehicle_id=str(policy_raw.get("forced_vehicle_id")) if policy_raw.get("forced_vehicle_id") else None,
        max_vehicles=int(policy_raw.get("max_vehicles", 5)),
    )
    selected = catalog
    if policy.mode == "forced":
        try:
            selected = (find_vehicle(policy.forced_vehicle_id or "", catalog),)
        except KeyError as exc:
            raise DomainError(
                Diagnostic("UNKNOWN_VEHICLE", f"Véhicule inconnu : {policy.forced_vehicle_id}")
            ) from exc
    return OptimizationProblem(
        items=tuple(expanded),
        vehicles=selected,
        vehicle_policy=policy,
        seed=int(payload.get("seed", 1)),
        budget_seconds=min(float(payload.get("budget_seconds", 30)), 30.0),
        requested_solutions=requested_solutions or int(payload.get("requested_solutions", 5)),
    )


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().replace("’", "'").lower()
    text = "".join(character for character in unicodedata.normalize("NFD", text) if not unicodedata.combining(character))
    return re.sub(r"\s+", " ", text)


COLUMN_ALIASES = {
    "id": {"id", "identifiant", "reference", "référence"},
    "quantity": {"quantity", "quantite", "quantité", "qty", "qté"},
    "shape": {"shape", "forme", "type"},
    "length": {"length", "longueur", "longueur (mm)", "longueur (cm)", "longueur (m)"},
    "width": {"width", "largeur", "largeur (mm)", "largeur (cm)", "largeur (m)"},
    "height": {"height", "hauteur", "h (mm)", "h (cm)", "h (m)", "hauteur (mm)", "hauteur (cm)", "hauteur (m)"},
    "weight": {"weight", "poids", "poids (kg)"},
    "destination": {"destination", "client"},
    "pickup_address": {"pickup_address", "point d’enlèvement", "point d'enlèvement", "enlèvement", "enlevement"},
    "delivery_address": {"delivery_address", "point de livraison", "livraison"},
    "delivery_order": {"delivery_order", "ordre_livraison", "ordre"},
    "rotation_allowed": {"rotation_allowed", "rotation_autorisee", "rotation_autorisée", "rotation autorisée"},
    "stackable": {"stackable", "gerbable"},
    "keep_together_group": {"keep_together_group", "groupe ensemble"},
    "separate_group": {"separate_group", "groupe séparé", "groupe separe"},
    "compatibility_tags": {"compatibility_tags", "tags compatibles"},
    "incompatible_tags": {"incompatible_tags", "tags incompatibles"},
    "separation": {
        "separation",
        "séparation",
        "séparation (mm)",
        "separation (mm)",
        "séparation (cm)",
        "separation (cm)",
        "séparation (m)",
        "separation (m)",
    },
}
_NORMALIZED_ALIASES = {
    target: {_normalize_header(alias) for alias in aliases}
    for target, aliases in COLUMN_ALIASES.items()
}


def _canonical_row(row: Mapping[str, Any]) -> dict[str, Any]:
    canonical: dict[str, Any] = {}
    exact = {str(key).strip(): value for key, value in row.items() if key is not None}
    for key, value in exact.items():
        compact = re.sub(r"\s+", "", key)
        if re.fullmatch(r"L\((mm|cm|m)\)", compact):
            canonical["length"] = value
        elif re.fullmatch(r"l\((mm|cm|m)\)", compact):
            canonical["width"] = value
    lowered = {_normalize_header(key): value for key, value in exact.items()}
    for target, aliases in _NORMALIZED_ALIASES.items():
        if target in canonical:
            continue
        for alias in aliases:
            if alias in lowered:
                canonical[target] = lowered[alias]
                break
    known_aliases = {alias for aliases in _NORMALIZED_ALIASES.values() for alias in aliases}
    for key, value in lowered.items():
        if key not in known_aliases:
            canonical.setdefault(key, value)
    return canonical


def _dimension_unit_from_headers(headers: list[str], fallback: str = "mm") -> str:
    units: set[str] = set()
    for header in headers:
        normalized = _normalize_header(header)
        if not any(word in normalized for word in ("longueur", "largeur", "hauteur", "separation")) and not re.match(
            r"^[Llh]\s*\(", str(header).strip()
        ):
            continue
        match = re.search(r"\((mm|cm|m)\)", normalized)
        if match:
            units.add(match.group(1))
    if len(units) > 1:
        raise DomainError(
            Diagnostic(
                "MIXED_DIMENSION_UNITS",
                "Le fichier mélange plusieurs unités de dimensions dans ses en-têtes. Utilisez une seule unité pour L, l, H et Séparation.",
            )
        )
    unit = next(iter(units), str(fallback or "mm").lower())
    if unit not in _DIMENSION_FACTORS:
        raise DomainError(
            Diagnostic("INVALID_UNIT", "L’unité indiquée dans le fichier n’est pas reconnue. Utilisez mm, cm ou m.")
        )
    return unit


def _validate_headers(headers: list[str]) -> None:
    mapped = set(_canonical_row({header: "" for header in headers}))
    missing = sorted(_REQUIRED_IMPORT_COLUMNS - mapped)
    if not missing:
        return
    labels = {
        "length": "L (mm) ou Longueur",
        "width": "l (mm) ou Largeur",
        "height": "H (mm) ou Hauteur",
        "weight": "Poids (kg)",
    }
    raise DomainError(
        Diagnostic(
            "MISSING_COLUMNS",
            "Colonnes obligatoires absentes : " + ", ".join(labels[field] for field in missing) + ". Téléchargez le modèle AxioLoad pour reprendre les intitulés attendus.",
            field_path="headers",
        )
    )


def payload_from_csv(content: bytes, **base: Any) -> dict[str, Any]:
    if not content:
        raise DomainError(Diagnostic("EMPTY_FILE", "Le fichier CSV est vide."))
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise DomainError(
                Diagnostic(
                    "INVALID_CSV_ENCODING",
                    "Le fichier CSV n’utilise pas un encodage lisible. Enregistrez-le en CSV UTF-8 depuis Excel.",
                )
            ) from exc
    if "\x00" in text:
        raise DomainError(
            Diagnostic(
                "INVALID_CSV_CONTENT",
                "Le fichier sélectionné n’est pas un CSV texte valide. Il s’agit peut-être d’un fichier Excel renommé en .csv.",
            )
        )
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        delimiter = ";" if text.count(";") >= text.count(",") else ","
        dialect = csv.excel
        dialect.delimiter = delimiter
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [str(header).strip() for header in (reader.fieldnames or []) if header is not None]
    if not headers:
        raise DomainError(
            Diagnostic(
                "MISSING_HEADER_ROW",
                "Aucune ligne d’en-tête n’a été trouvée dans le CSV. La première ligne doit contenir les noms de colonnes.",
            )
        )
    _validate_headers(headers)
    dimension_unit = _dimension_unit_from_headers(headers, str(base.get("dimension_unit", "mm")))
    items: list[dict[str, Any]] = []
    for source_row, row in enumerate(reader, start=2):
        if not any(value not in (None, "") for value in row.values()):
            continue
        item = _canonical_row(row)
        item["_source_row"] = source_row
        items.append(item)
    if not items:
        raise DomainError(
            Diagnostic(
                "EMPTY_DATA_ROWS",
                "Le CSV contient des en-têtes mais aucune ligne de marchandise renseignée.",
            )
        )
    return {**base, "dimension_unit": dimension_unit, "items": items}


def payload_from_xlsx(content: bytes, **base: Any) -> dict[str, Any]:
    if not content:
        raise DomainError(Diagnostic("EMPTY_FILE", "Le fichier Excel est vide."))
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise DomainError(
            Diagnostic(
                "INVALID_XLSX_FILE",
                "Le fichier n’est pas un classeur XLSX valide ou il est endommagé. Ouvrez-le dans Excel ou LibreOffice, puis utilisez « Enregistrer sous » au format .xlsx.",
            )
        ) from exc

    best: tuple[int, int, Any, list[tuple[Any, ...]]] | None = None
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        for header_index, row in enumerate(rows[:20]):
            headers = [str(value).strip() if value is not None else "" for value in row]
            mapped = set(_canonical_row({header: "" for header in headers if header}))
            score = len(mapped & _REQUIRED_IMPORT_COLUMNS)
            candidate = (score, -header_index, sheet, rows)
            if best is None or candidate[:2] > best[:2]:
                best = candidate

    if best is None or best[0] == 0:
        raise DomainError(
            Diagnostic(
                "MISSING_HEADER_ROW",
                "Aucune feuille ne contient une ligne d’en-tête reconnue. Les en-têtes doivent se trouver dans les 20 premières lignes.",
            )
        )

    _, negative_header_index, sheet, rows = best
    header_index = -negative_header_index
    headers = [str(value).strip() if value is not None else "" for value in rows[header_index]]
    _validate_headers(headers)
    dimension_unit = _dimension_unit_from_headers(headers, str(base.get("dimension_unit", "mm")))
    items: list[dict[str, Any]] = []
    for source_row, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(value is not None and str(value).strip() != "" for value in row):
            continue
        item = _canonical_row(dict(zip(headers, row, strict=False)))
        item["_source_row"] = source_row
        items.append(item)
    if not items:
        raise DomainError(
            Diagnostic(
                "EMPTY_DATA_ROWS",
                f"La feuille « {sheet.title} » contient des en-têtes mais aucune ligne de marchandise renseignée.",
            )
        )
    return {
        **base,
        "dimension_unit": dimension_unit,
        "items": items,
        "_import_sheet": sheet.title,
        "_import_header_row": header_index + 1,
    }
